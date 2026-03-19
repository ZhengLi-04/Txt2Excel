#!/usr/bin/env python
"""
Txt Excel Studio - 统一的 TXT 转 Excel 小工具

覆盖原有三类处理方式：
1. 通用分列：按分隔符拆分并横向/纵向合并到 Excel。
2. CHI 数据：提取表头后的数据，可按最后 N 行、最后 N 秒或时间范围筛选。
3. 原始文本：保留原始行，可选只取最后 N 行。
"""

from __future__ import annotations

import os
import csv
from collections import deque
from dataclasses import dataclass
from typing import Iterable

import tkinter as tk
from tkinter import filedialog, messagebox, ttk

from openpyxl import Workbook


ENCODINGS = ("utf-8", "gbk", "latin-1")
LAYOUT_OPTIONS = {
    "horizontal": "横向拼接",
    "vertical": "纵向合并",
    "both": "同时输出横向和纵向",
}
PROFILE_OPTIONS = {
    "generic": "通用分列",
    "chi": "CHI 电化学数据",
    "raw": "原始文本/末尾行",
}


@dataclass
class Dataset:
    title: str
    rows: list[list[str]]

    @property
    def width(self) -> int:
        return max((len(row) for row in self.rows), default=1)


def coerce_cell_value(value: str):
    stripped = value.strip()
    if not stripped:
        return value

    integer_candidate = stripped.replace(",", "")
    if integer_candidate.isdigit() or (
        integer_candidate.startswith("-") and integer_candidate[1:].isdigit()
    ):
        try:
            return int(integer_candidate)
        except ValueError:
            pass

    try:
        return float(stripped)
    except ValueError:
        return value


def read_text_lines(filepath: str) -> list[str]:
    last_error = None
    for encoding in ENCODINGS:
        try:
            with open(filepath, "r", encoding=encoding) as handle:
                return [line.rstrip("\r\n") for line in handle]
        except UnicodeDecodeError as exc:
            last_error = exc
    raise last_error or RuntimeError(f"无法读取文件: {filepath}")


def read_last_n_lines(filepath: str, count: int) -> list[str]:
    if count <= 0:
        return read_text_lines(filepath)

    last_error = None
    for encoding in ENCODINGS:
        try:
            with open(filepath, "r", encoding=encoding) as handle:
                return [line.rstrip("\r\n") for line in deque(handle, maxlen=count)]
        except UnicodeDecodeError as exc:
            last_error = exc
    raise last_error or RuntimeError(f"无法读取文件: {filepath}")


def split_lines(lines: Iterable[str], delimiter: str, skip_empty: bool) -> list[list[str]]:
    rows: list[list[str]] = []
    for line in lines:
        stripped = line.strip()
        if not stripped and skip_empty:
            continue
        if delimiter:
            rows.append([cell.strip() for cell in line.split(delimiter)])
        else:
            rows.append([line])
    return rows


def to_float(value: str) -> float | None:
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def parse_chi_file(filepath: str, header_prefix: str) -> tuple[str, list[list[str]]]:
    file_name = os.path.basename(filepath)
    rows: list[list[str]] = []
    header_found = False

    for line in read_text_lines(filepath):
        stripped = line.strip()
        if stripped.startswith("File:"):
            file_name = stripped.split("File:", 1)[1].strip() or file_name
        if not header_found and stripped.startswith(header_prefix):
            header_found = True
            rows.append([cell.strip() for cell in stripped.split(",")])
            continue
        if header_found and stripped:
            rows.append([cell.strip() for cell in stripped.split(",")])

    if not rows:
        raise ValueError(f"未在文件中找到 CHI 数据表头: {os.path.basename(filepath)}")
    return file_name, rows


def filter_chi_rows(
    rows: list[list[str]],
    extraction_mode: str,
    last_count: int,
    last_seconds: float,
    time_min: float | None,
    time_max: float | None,
) -> list[list[str]]:
    if not rows:
        return rows

    header, data_rows = rows[0], rows[1:]
    if not data_rows:
        return rows

    if extraction_mode == "all":
        filtered = data_rows
    elif extraction_mode == "last_points":
        filtered = data_rows[-last_count:] if last_count > 0 else data_rows
    elif extraction_mode == "last_seconds":
        if last_seconds <= 0:
            filtered = data_rows
        else:
            last_time = next((to_float(row[0]) for row in reversed(data_rows) if row), None)
            if last_time is None:
                raise ValueError("CHI 数据第一列不是可识别的时间数值")
            threshold = last_time - last_seconds
            filtered = [row for row in data_rows if row and (to_float(row[0]) or 0) >= threshold]
    elif extraction_mode == "time_range":
        filtered = []
        for row in data_rows:
            if not row:
                continue
            time_value = to_float(row[0])
            if time_value is None:
                continue
            if time_min is not None and time_value < time_min:
                continue
            if time_max is not None and time_value > time_max:
                continue
            filtered.append(row)
    else:
        raise ValueError(f"未知 CHI 提取模式: {extraction_mode}")

    return [header] + filtered


def build_datasets(options: dict) -> list[Dataset]:
    profile = options["profile"]
    datasets: list[Dataset] = []

    for filepath in options["files"]:
        base_name = os.path.basename(filepath)
        if profile == "chi":
            title, rows = parse_chi_file(filepath, options["chi_header_prefix"])
            rows = filter_chi_rows(
                rows,
                options["chi_extract_mode"],
                options["chi_last_points"],
                options["chi_last_seconds"],
                options["time_min"],
                options["time_max"],
            )
        elif profile == "raw":
            lines = read_last_n_lines(filepath, options["raw_last_lines"])
            if options["skip_empty"]:
                lines = [line for line in lines if line.strip()]
            title = base_name
            rows = [[line] for line in lines]
        else:
            lines = read_text_lines(filepath)
            if options["generic_last_lines"] > 0:
                lines = lines[-options["generic_last_lines"] :]
            title = base_name
            rows = split_lines(lines, options["delimiter"], options["skip_empty"])

        datasets.append(Dataset(title=title, rows=rows))

    if not datasets:
        raise ValueError("未生成任何可导出的数据")
    return datasets


def write_horizontal_sheet(worksheet, datasets: list[Dataset], include_title: bool, title_gap: int, col_gap: int) -> None:
    current_col = 1
    start_row = 1

    for dataset in datasets:
        data_start_row = start_row
        if include_title:
            worksheet.cell(row=start_row, column=current_col, value=dataset.title)
            data_start_row += title_gap + 1

        for row_index, row_data in enumerate(dataset.rows, start=data_start_row):
            for col_index, cell_value in enumerate(row_data, start=0):
                worksheet.cell(row=row_index, column=current_col + col_index, value=coerce_cell_value(cell_value))

        current_col += dataset.width + col_gap


def write_vertical_sheet(worksheet, datasets: list[Dataset], include_title: bool, title_gap: int, row_gap: int) -> None:
    current_row = 1

    for dataset in datasets:
        if include_title:
            worksheet.cell(row=current_row, column=1, value=dataset.title)
            current_row += title_gap + 1

        for row_data in dataset.rows:
            for col_index, cell_value in enumerate(row_data, start=1):
                worksheet.cell(row=current_row, column=col_index, value=coerce_cell_value(cell_value))
            current_row += 1

        current_row += row_gap


def datasets_to_rows(datasets: list[Dataset], layout: str, include_title: bool, title_gap: int, col_gap: int, row_gap: int) -> list[list[object]]:
    if layout == "vertical":
        output_rows: list[list[object]] = []
        for dataset in datasets:
            if include_title:
                output_rows.append([dataset.title])
                for _ in range(title_gap):
                    output_rows.append([])
            for row_data in dataset.rows:
                output_rows.append([coerce_cell_value(cell) for cell in row_data])
            for _ in range(row_gap):
                output_rows.append([])
        return output_rows

    if layout == "horizontal":
        grid: list[list[object]] = []
        current_col = 0
        start_row = 0

        for dataset in datasets:
            data_start_row = start_row
            if include_title:
                while len(grid) <= start_row:
                    grid.append([])
                row = grid[start_row]
                while len(row) <= current_col:
                    row.append("")
                row[current_col] = dataset.title
                data_start_row += title_gap + 1

            for row_offset, row_data in enumerate(dataset.rows):
                target_row = data_start_row + row_offset
                while len(grid) <= target_row:
                    grid.append([])
                row = grid[target_row]
                needed_width = current_col + len(row_data)
                while len(row) < needed_width:
                    row.append("")
                for col_offset, cell_value in enumerate(row_data):
                    row[current_col + col_offset] = coerce_cell_value(cell_value)

            current_col += dataset.width + col_gap

        return grid

    raise ValueError("CSV 仅支持横向拼接或纵向合并")


def export_to_csv(datasets: list[Dataset], save_path: str, layout: str, include_title: bool, title_gap: int, col_gap: int, row_gap: int) -> None:
    rows = datasets_to_rows(datasets, layout, include_title, title_gap, col_gap, row_gap)
    with open(save_path, "w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.writer(handle)
        writer.writerows(rows)


def export_to_excel(datasets: list[Dataset], save_path: str, layout: str, include_title: bool, title_gap: int, col_gap: int, row_gap: int) -> None:
    workbook = Workbook()

    if layout == "horizontal":
        sheet = workbook.active
        sheet.title = "Horizontal Merge"
        write_horizontal_sheet(sheet, datasets, include_title, title_gap, col_gap)
    elif layout == "vertical":
        sheet = workbook.active
        sheet.title = "Vertical Merge"
        write_vertical_sheet(sheet, datasets, include_title, title_gap, row_gap)
    elif layout == "both":
        horizontal_sheet = workbook.active
        horizontal_sheet.title = "Horizontal Merge"
        vertical_sheet = workbook.create_sheet("Vertical Merge")
        write_horizontal_sheet(horizontal_sheet, datasets, include_title, title_gap, col_gap)
        write_vertical_sheet(vertical_sheet, datasets, include_title, title_gap, row_gap)
    else:
        raise ValueError(f"未知输出布局: {layout}")

    workbook.save(save_path)


def export_data(datasets: list[Dataset], save_path: str, layout: str, include_title: bool, title_gap: int, col_gap: int, row_gap: int) -> None:
    extension = os.path.splitext(save_path)[1].lower()
    if extension == ".csv":
        if layout == "both":
            raise ValueError("CSV 不能同时导出双表，请选择横向拼接或纵向合并")
        export_to_csv(datasets, save_path, layout, include_title, title_gap, col_gap, row_gap)
        return

    export_to_excel(datasets, save_path, layout, include_title, title_gap, col_gap, row_gap)


class TxtExcelStudioApp:
    def __init__(self, default_profile: str = "generic") -> None:
        self.root = tk.Tk()
        self.root.title("Txt Excel Studio")
        self.root.geometry("900x720")

        self.files: list[str] = []

        self.profile_var = tk.StringVar(value=default_profile)
        self.layout_var = tk.StringVar(value="both" if default_profile == "chi" else "horizontal")
        self.include_title_var = tk.BooleanVar(value=True)
        self.skip_empty_var = tk.BooleanVar(value=True)
        self.delimiter_var = tk.StringVar(value=",")
        self.title_gap_var = tk.IntVar(value=1)
        self.col_gap_var = tk.IntVar(value=2)
        self.row_gap_var = tk.IntVar(value=1)
        self.generic_last_lines_var = tk.IntVar(value=0)
        self.raw_last_lines_var = tk.IntVar(value=300)
        self.chi_header_prefix_var = tk.StringVar(value="Time/s")
        self.chi_extract_mode_var = tk.StringVar(value="last_points")
        self.chi_last_points_var = tk.IntVar(value=300)
        self.chi_last_seconds_var = tk.DoubleVar(value=30.0)
        self.time_min_var = tk.StringVar(value="")
        self.time_max_var = tk.StringVar(value="")

        self._build_ui()
        self._apply_profile_defaults(default_profile)
        self._update_option_visibility()

    def _build_ui(self) -> None:
        main = ttk.Frame(self.root, padding=12)
        main.pack(fill="both", expand=True)

        file_frame = ttk.LabelFrame(main, text="文件")
        file_frame.pack(fill="both", expand=False)

        button_row = ttk.Frame(file_frame)
        button_row.pack(fill="x", padx=8, pady=8)
        ttk.Button(button_row, text="添加文件", command=self.add_files).pack(side="left")
        ttk.Button(button_row, text="移除选中", command=self.remove_selected_files).pack(side="left", padx=(8, 0))
        ttk.Button(button_row, text="上移", command=lambda: self.move_selected(-1)).pack(side="left", padx=(8, 0))
        ttk.Button(button_row, text="下移", command=lambda: self.move_selected(1)).pack(side="left", padx=(8, 0))
        ttk.Button(button_row, text="清空", command=self.clear_files).pack(side="left", padx=(8, 0))

        self.file_listbox = tk.Listbox(file_frame, height=10, selectmode=tk.EXTENDED)
        self.file_listbox.pack(fill="both", expand=True, padx=8, pady=(0, 8))

        options_frame = ttk.LabelFrame(main, text="处理选项")
        options_frame.pack(fill="x", expand=False, pady=(12, 0))
        options_frame.columnconfigure(1, weight=1)

        ttk.Label(options_frame, text="处理模式").grid(row=0, column=0, sticky="w", padx=8, pady=6)
        profile_combo = ttk.Combobox(
            options_frame,
            textvariable=self.profile_var,
            state="readonly",
            values=list(PROFILE_OPTIONS.keys()),
        )
        profile_combo.grid(row=0, column=1, sticky="ew", padx=8, pady=6)
        profile_combo.bind("<<ComboboxSelected>>", lambda _event: self.on_profile_changed())

        self.profile_hint = ttk.Label(options_frame, text="")
        self.profile_hint.grid(row=0, column=2, sticky="w", padx=8, pady=6)

        ttk.Label(options_frame, text="输出布局").grid(row=1, column=0, sticky="w", padx=8, pady=6)
        layout_combo = ttk.Combobox(
            options_frame,
            textvariable=self.layout_var,
            state="readonly",
            values=list(LAYOUT_OPTIONS.keys()),
        )
        layout_combo.grid(row=1, column=1, sticky="ew", padx=8, pady=6)
        self.layout_hint = ttk.Label(options_frame, text="")
        self.layout_hint.grid(row=1, column=2, sticky="w", padx=8, pady=6)

        ttk.Checkbutton(options_frame, text="在每组数据前写入文件名", variable=self.include_title_var).grid(
            row=2, column=0, columnspan=2, sticky="w", padx=8, pady=6
        )
        ttk.Checkbutton(options_frame, text="跳过空行", variable=self.skip_empty_var).grid(
            row=2, column=2, sticky="w", padx=8, pady=6
        )

        ttk.Label(options_frame, text="标题后空行数").grid(row=3, column=0, sticky="w", padx=8, pady=6)
        ttk.Entry(options_frame, textvariable=self.title_gap_var, width=10).grid(row=3, column=1, sticky="w", padx=8, pady=6)

        ttk.Label(options_frame, text="横向空列数").grid(row=4, column=0, sticky="w", padx=8, pady=6)
        ttk.Entry(options_frame, textvariable=self.col_gap_var, width=10).grid(row=4, column=1, sticky="w", padx=8, pady=6)

        ttk.Label(options_frame, text="纵向空行数").grid(row=5, column=0, sticky="w", padx=8, pady=6)
        ttk.Entry(options_frame, textvariable=self.row_gap_var, width=10).grid(row=5, column=1, sticky="w", padx=8, pady=6)

        self.generic_frame = ttk.LabelFrame(main, text="通用分列选项")
        self.generic_frame.pack(fill="x", expand=False, pady=(12, 0))
        ttk.Label(self.generic_frame, text="分隔符").grid(row=0, column=0, sticky="w", padx=8, pady=6)
        ttk.Entry(self.generic_frame, textvariable=self.delimiter_var, width=10).grid(row=0, column=1, sticky="w", padx=8, pady=6)
        ttk.Label(self.generic_frame, text="只取最后 N 行，0 表示全部").grid(row=1, column=0, sticky="w", padx=8, pady=6)
        ttk.Entry(self.generic_frame, textvariable=self.generic_last_lines_var, width=10).grid(
            row=1, column=1, sticky="w", padx=8, pady=6
        )

        self.raw_frame = ttk.LabelFrame(main, text="原始文本选项")
        self.raw_frame.pack(fill="x", expand=False, pady=(12, 0))
        ttk.Label(self.raw_frame, text="保留最后 N 行，0 表示全部").grid(row=0, column=0, sticky="w", padx=8, pady=6)
        ttk.Entry(self.raw_frame, textvariable=self.raw_last_lines_var, width=10).grid(row=0, column=1, sticky="w", padx=8, pady=6)

        self.chi_frame = ttk.LabelFrame(main, text="CHI 选项")
        self.chi_frame.pack(fill="x", expand=False, pady=(12, 0))
        self.chi_frame.columnconfigure(1, weight=1)

        ttk.Label(self.chi_frame, text="数据表头前缀").grid(row=0, column=0, sticky="w", padx=8, pady=6)
        ttk.Entry(self.chi_frame, textvariable=self.chi_header_prefix_var, width=20).grid(
            row=0, column=1, sticky="w", padx=8, pady=6
        )

        ttk.Label(self.chi_frame, text="提取方式").grid(row=1, column=0, sticky="w", padx=8, pady=6)
        chi_mode = ttk.Combobox(
            self.chi_frame,
            textvariable=self.chi_extract_mode_var,
            state="readonly",
            values=("all", "last_points", "last_seconds", "time_range"),
        )
        chi_mode.grid(row=1, column=1, sticky="ew", padx=8, pady=6)
        chi_mode.bind("<<ComboboxSelected>>", lambda _event: self._update_option_visibility())

        self.chi_last_points_label = ttk.Label(self.chi_frame, text="最后 N 行")
        self.chi_last_points_entry = ttk.Entry(self.chi_frame, textvariable=self.chi_last_points_var, width=12)
        self.chi_last_seconds_label = ttk.Label(self.chi_frame, text="最后 N 秒")
        self.chi_last_seconds_entry = ttk.Entry(self.chi_frame, textvariable=self.chi_last_seconds_var, width=12)
        self.time_min_label = ttk.Label(self.chi_frame, text="起始时间")
        self.time_min_entry = ttk.Entry(self.chi_frame, textvariable=self.time_min_var, width=12)
        self.time_max_label = ttk.Label(self.chi_frame, text="结束时间")
        self.time_max_entry = ttk.Entry(self.chi_frame, textvariable=self.time_max_var, width=12)

        action_frame = ttk.Frame(main)
        action_frame.pack(fill="x", expand=False, pady=(16, 0))
        ttk.Button(action_frame, text="导出 Excel", command=self.export).pack(side="left")
        ttk.Button(action_frame, text="退出", command=self.root.destroy).pack(side="right")

    def _set_listbox_items(self) -> None:
        self.file_listbox.delete(0, tk.END)
        for path in self.files:
            self.file_listbox.insert(tk.END, path)

    def add_files(self) -> None:
        selected = filedialog.askopenfilenames(
            title="选择要处理的 TXT 文件",
            filetypes=[("Text files", "*.txt"), ("All files", "*.*")],
        )
        if not selected:
            return
        self.files.extend(path for path in selected if path not in self.files)
        self._set_listbox_items()

    def remove_selected_files(self) -> None:
        indexes = list(self.file_listbox.curselection())
        if not indexes:
            return
        for index in reversed(indexes):
            del self.files[index]
        self._set_listbox_items()

    def clear_files(self) -> None:
        self.files.clear()
        self._set_listbox_items()

    def move_selected(self, offset: int) -> None:
        indexes = list(self.file_listbox.curselection())
        if len(indexes) != 1:
            return

        index = indexes[0]
        new_index = index + offset
        if new_index < 0 or new_index >= len(self.files):
            return

        self.files[index], self.files[new_index] = self.files[new_index], self.files[index]
        self._set_listbox_items()
        self.file_listbox.selection_set(new_index)

    def _apply_profile_defaults(self, profile: str) -> None:
        if profile == "chi":
            self.layout_var.set("both")
            self.include_title_var.set(True)
            self.chi_extract_mode_var.set("last_points")
            self.chi_last_points_var.set(300)
            self.chi_last_seconds_var.set(30.0)
        elif profile == "raw":
            self.layout_var.set("vertical")
            self.include_title_var.set(False)
            self.raw_last_lines_var.set(300)
        else:
            self.layout_var.set("horizontal")
            self.include_title_var.set(True)
            self.delimiter_var.set(",")

    def on_profile_changed(self) -> None:
        self._apply_profile_defaults(self.profile_var.get())
        self._update_option_visibility()

    def _update_option_visibility(self) -> None:
        profile = self.profile_var.get()
        self.profile_hint.configure(text=PROFILE_OPTIONS.get(profile, ""))
        self.layout_hint.configure(text=LAYOUT_OPTIONS.get(self.layout_var.get(), ""))

        for frame, active in (
            (self.generic_frame, profile == "generic"),
            (self.raw_frame, profile == "raw"),
            (self.chi_frame, profile == "chi"),
        ):
            if active:
                frame.pack(fill="x", expand=False, pady=(12, 0))
            else:
                frame.pack_forget()

        for widget in (
            self.chi_last_points_label,
            self.chi_last_points_entry,
            self.chi_last_seconds_label,
            self.chi_last_seconds_entry,
            self.time_min_label,
            self.time_min_entry,
            self.time_max_label,
            self.time_max_entry,
        ):
            widget.grid_forget()

        chi_mode = self.chi_extract_mode_var.get()
        if profile != "chi":
            return
        if chi_mode == "last_points":
            self.chi_last_points_label.grid(row=2, column=0, sticky="w", padx=8, pady=6)
            self.chi_last_points_entry.grid(row=2, column=1, sticky="w", padx=8, pady=6)
        elif chi_mode == "last_seconds":
            self.chi_last_seconds_label.grid(row=2, column=0, sticky="w", padx=8, pady=6)
            self.chi_last_seconds_entry.grid(row=2, column=1, sticky="w", padx=8, pady=6)
        elif chi_mode == "time_range":
            self.time_min_label.grid(row=2, column=0, sticky="w", padx=8, pady=6)
            self.time_min_entry.grid(row=2, column=1, sticky="w", padx=8, pady=6)
            self.time_max_label.grid(row=3, column=0, sticky="w", padx=8, pady=6)
            self.time_max_entry.grid(row=3, column=1, sticky="w", padx=8, pady=6)

    def _collect_options(self) -> dict:
        if not self.files:
            raise ValueError("请先添加至少一个 TXT 文件")

        time_min = self.time_min_var.get().strip()
        time_max = self.time_max_var.get().strip()
        return {
            "files": self.files,
            "profile": self.profile_var.get(),
            "layout": self.layout_var.get(),
            "include_title": self.include_title_var.get(),
            "skip_empty": self.skip_empty_var.get(),
            "delimiter": self.delimiter_var.get(),
            "title_gap": max(0, self.title_gap_var.get()),
            "col_gap": max(0, self.col_gap_var.get()),
            "row_gap": max(0, self.row_gap_var.get()),
            "generic_last_lines": max(0, self.generic_last_lines_var.get()),
            "raw_last_lines": max(0, self.raw_last_lines_var.get()),
            "chi_header_prefix": self.chi_header_prefix_var.get().strip() or "Time/s",
            "chi_extract_mode": self.chi_extract_mode_var.get(),
            "chi_last_points": max(0, self.chi_last_points_var.get()),
            "chi_last_seconds": max(0.0, self.chi_last_seconds_var.get()),
            "time_min": float(time_min) if time_min else None,
            "time_max": float(time_max) if time_max else None,
        }

    def export(self) -> None:
        try:
            options = self._collect_options()
            datasets = build_datasets(options)
        except Exception as exc:
            messagebox.showerror("处理失败", str(exc))
            return

        save_path = filedialog.asksaveasfilename(
            title="保存导出文件",
            defaultextension=".xlsx",
            filetypes=[
                ("Excel files", "*.xlsx"),
                ("CSV files", "*.csv"),
                ("All files", "*.*"),
            ],
        )
        if not save_path:
            return

        try:
            export_data(
                datasets,
                save_path,
                options["layout"],
                options["include_title"],
                options["title_gap"],
                options["col_gap"],
                options["row_gap"],
            )
        except Exception as exc:
            messagebox.showerror("导出失败", str(exc))
            return

        messagebox.showinfo("完成", f"Excel 已导出到:\n{save_path}")

    def run(self) -> None:
        self.root.mainloop()


def launch_app(default_profile: str = "generic") -> None:
    app = TxtExcelStudioApp(default_profile=default_profile)
    app.run()


def main() -> None:
    launch_app()


if __name__ == "__main__":
    main()
